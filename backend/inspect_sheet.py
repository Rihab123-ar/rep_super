from pathlib import Path
from openpyxl import load_workbook

file_path = Path(__file__).parent / 'data' / 'sinistres.xlsx'
print('Inspecting', file_path)
if not file_path.exists():
    raise SystemExit('File not found: ' + str(file_path))

wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
print('sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames[:2]:
    ws = wb[sheet_name]
    print('--- Sheet:', sheet_name)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        print(f'row {i}:', row)
    print('')

# Count is_fraud values if present
ws = wb.active
header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
if 'is_fraud' in header:
    idx = header.index('is_fraud')
    from collections import Counter
    counts = Counter()
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        counts[row[idx]] += 1
        total += 1
    print('is_fraud counts:', dict(counts))
    print('total rows counted:', total)
else:
    print('is_fraud column not found in active sheet')
wb.close()
